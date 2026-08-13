from __future__ import annotations

from io import BytesIO
import pymupdf
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from sublist import build_sublist_pdf, parse_packing_list


st.set_page_config(page_title="Packing List → Sublist A5", page_icon="📦", layout="centered")
st.title("Packing List → Sublist A5")
st.caption("Tạo một trang sublist A5 cho mỗi carton, sẵn sàng để in ở 100% / Actual size.")

def make_sample_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Packing List"
    sheet.append(["PACKING LIST（装箱单）"])
    sheet.append(["WPIC Purchase Order#/箱单编号：", None, None, None, None, None, None, None, None, None, None, "日期/Date："])
    sheet.append(["Seller's EIN#：", None, None, None, None, None, None, None, None, None, None, "Invoice#:"])
    sheet.append(["SHIPPER:\nSAMPLE COMPANY\nSample address", None, None, None, None, None, None, None, None, None, None, "Remark (SO#):"])
    sheet.append(["CONSIGNEE:\nSAMPLE CUSTOMER\nSample delivery address", None, None, None, None, None, None, None, None, None, None, "NOTIFY PARTY:\nSAMPLE CUSTOMER"])
    sheet.append(["成交方式/Trade term："])
    sheet.append([])
    sheet.append(["Package Total:", "=J18", "2 Cartons"])
    sheet.append(["Quantity Total:", "=I18", 4])
    sheet.append(["Gross Weight (KG):", "=O18", 1.05])
    sheet.append(["CBM", "=P18", 0.02])
    sheet.append([
        "Item#", "Store", "OR No.", "Ref No.", "Product Name\nin English", "SKU#",
        "BarCode/UPC", "UOM", "Quantity", "Carton#", "Packaging code",
        "Carton Dimensions (cm)\n(Length*Width*Height)", None, None, "Weight (KG)",
        "CBM", "Origin Country", "Origin Country's HTSCODE", "Shipping Mark", "PORT", "中国标签名称",
    ])
    sheet.append(["项目", None, "OR 编码", None, "货品名称", "SKU编码", "条形码", "单位", "数量", "箱号", "包装条形码", "箱子尺寸", None, None, None, None, "原产国", "原产国"])
    sheet.append([1, None, "OR-SAMPLE", "REF001", "Sample Product A", "SKU-SAMPLE-01", "4890000000001", "PCS", 1, "1/2", "PKG-SAMPLE-0001", 26, 18, 11, 0.45, 0.01, "VN", "4202.92.31", "SAMPLE_VN"])
    sheet.append([2, None, "OR-SAMPLE", "REF001", "Sample Product B", "SKU-SAMPLE-02", "4890000000002", "PCS", 1, "2/2", "PKG-SAMPLE-0002", 26, 18, 11, 0.60, 0.01, "CN", "5609.00.30", "SAMPLE_CN"])
    sheet.append([3, None, "OR-SAMPLE", "REF001", "Sample Product C", "SKU-SAMPLE-03", "4890000000003", "PCS", 1, None, None, None, None, None, None, None, "CN", "5609.00.30", "SAMPLE_CN"])
    sheet.append([4, None, "OR-SAMPLE", "REF001", "Sample Product D", "SKU-SAMPLE-04", "4890000000004", "PCS", 1, None, None, None, None, None, None, None, "CN", "5609.00.30", "SAMPLE_CN"])
    sheet.append(["TOTAL", None, None, None, None, None, None, None, 4, "2 Cartons", None, None, None, None, 1.05, 0.02])

    sheet.merge_cells("A1:U1")
    for row in (12, 13):
        for cell in sheet[row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet["A1"].font = Font(bold=True, size=16)
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.row_dimensions[1].height = 26
    sheet.row_dimensions[12].height = 42
    sheet.row_dimensions[13].height = 28
    widths = [10, 12, 14, 14, 24, 25, 18, 10, 12, 12, 24, 16, 10, 10, 14, 12, 16, 22, 18, 12, 18]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A14"
    sheet.auto_filter.ref = "A12:U18"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


st.download_button(
    "Tải file Excel mẫu",
    make_sample_workbook(),
    file_name="packing_list_mau.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

uploaded = st.file_uploader("Tải packing list Excel", type=["xlsx"])
if uploaded:
    try:
        cartons = parse_packing_list(BytesIO(uploaded.getvalue()))
        pdf_bytes = build_sublist_pdf(cartons)
    except Exception as exc:
        st.error(f"Không thể xử lý file: {exc}")
        st.stop()

    st.success(f"Đã tạo {len(cartons)} carton / {len(cartons)} trang A5.")
    st.download_button(
        "Tải xuống Sublist PDF A5",
        pdf_bytes,
        file_name=f"{uploaded.name.rsplit('.', 1)[0]}_sublist_A5.pdf",
        mime="application/pdf",
        type="primary",
    )

    st.subheader("Xem trước")
    doc = pymupdf.open(stream=pdf_bytes, filetype="pdf")
    for page_number, page in enumerate(doc, start=1):
        pix = page.get_pixmap(matrix=pymupdf.Matrix(1.35, 1.35), alpha=False)
        st.image(pix.tobytes("png"), caption=f"Carton {page_number}", use_container_width=True)
