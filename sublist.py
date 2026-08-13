from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal
from io import BytesIO
from typing import BinaryIO

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A5
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.platypus import Paragraph, Table, TableStyle
from reportlab.pdfgen import canvas


@dataclass
class Item:
    sku: str
    ean: str
    qty: str


@dataclass
class Carton:
    carton: str
    or_no: str
    ref_no: str
    gross_weight: str
    packaging_code: str
    items: list[Item] = field(default_factory=list)
    show_total: bool = True
    total_qty: str = ""


ALIASES = {
    "or_no": ("or no", "or #", "or code", "or 编码"),
    "ref_no": ("ref no", "ref #", "so no", "reference"),
    "sku": ("sku#", "sku编码", "style no", "item no"),
    "ean": ("barcode/upc", "barcode", "upc", "ean", "条形码"),
    "qty": ("quantity", "qty", "数量"),
    "carton": ("carton#", "carton no", "carton number", "箱号"),
    "packaging": ("packaging code", "packing code", "包装条形码"),
    "weight": ("weight (kg)", "gross weight", "gw", "重量"),
}

MAX_ITEMS_PER_PAGE = 25


def _normalized(value) -> str:
    return " ".join(str(value or "").replace("\n", " ").strip().lower().split())


def excel_text(value, number_format: str = "") -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        text = str(value)
    elif isinstance(value, float):
        text = decimal_text(Decimal(str(value)))
    else:
        return str(value).strip()
    if number_format and set(number_format) <= {"0"} and len(number_format) > len(text):
        return text.zfill(len(number_format))
    return text


def decimal_text(value: Decimal) -> str:
    """Format a decimal without ever removing significant integer zeroes."""
    text = format(value, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"


def _find_header(ws):
    for row in range(1, min(ws.max_row, 60) + 1):
        found = {}
        for cell in ws[row]:
            label = _normalized(cell.value)
            for key, aliases in ALIASES.items():
                if any(alias == label or alias in label for alias in aliases):
                    found.setdefault(key, cell.column)
        if all(key in found for key in ("sku", "ean", "qty", "carton", "packaging")):
            return row, found
    raise ValueError("Không tìm thấy dòng tiêu đề có SKU, EAN, Quantity, Carton# và Packaging code.")


def parse_packing_list(source: str | BinaryIO) -> list[Carton]:
    wb = load_workbook(source, data_only=True)
    candidates = []
    for ws in wb.worksheets:
        try:
            header_row, columns = _find_header(ws)
            candidates.append((ws, header_row, columns))
        except ValueError:
            continue
    if not candidates:
        raise ValueError("Không tìm thấy sheet packing list phù hợp.")
    ws, header_row, columns = max(candidates, key=lambda x: len(x[2]))

    cartons: list[Carton] = []
    current = None
    for row in range(header_row + 1, ws.max_row + 1):
        values = {key: ws.cell(row, col) for key, col in columns.items()}
        normalized_cells = [_normalized(cell.value) for cell in ws[row]]
        header_hits = sum(
            1 for aliases in ALIASES.values()
            if any(any(alias == cell or alias in cell for alias in aliases) for cell in normalized_cells)
        )
        if header_hits >= 3:
            continue
        sku = excel_text(values["sku"].value, values["sku"].number_format)
        label_blob = " ".join(normalized_cells)
        if any(token in label_blob for token in ("grand total", "total", "合计", "總計")):
            continue
        carton_value = excel_text(values["carton"].value, values["carton"].number_format)
        if carton_value:
            current = Carton(
                carton=carton_value,
                or_no=excel_text(values.get("or_no").value) if values.get("or_no") else "",
                ref_no=excel_text(values.get("ref_no").value) if values.get("ref_no") else "",
                gross_weight=excel_text(values.get("weight").value) if values.get("weight") else "",
                packaging_code=excel_text(values["packaging"].value, values["packaging"].number_format),
            )
            cartons.append(current)
        if not sku:
            continue
        if current is None:
            continue
        current.items.append(Item(
            sku=sku,
            ean=excel_text(values["ean"].value, values["ean"].number_format),
            qty=excel_text(values["qty"].value, values["qty"].number_format),
        ))
    if not cartons:
        raise ValueError("Không có carton hợp lệ trong file.")
    return cartons


def _fit_font(text: str, max_width: float, start: float, minimum: float = 6) -> float:
    size = start
    while size > minimum and stringWidth(text, "Helvetica", size) > max_width:
        size -= 0.25
    return size


def _qty_number(value: str):
    try:
        return Decimal(value)
    except Exception:
        return Decimal(0)


def _continued_carton_number(carton_number: str, part: int) -> str:
    """Turn 15/48 into 15-1/48 while preserving other carton formats."""
    if "/" in carton_number:
        current, total = carton_number.split("/", 1)
        return f"{current}-{part}/{total}"
    return f"{carton_number}-{part}"


def paginate_cartons(cartons: list[Carton], page_size: int = MAX_ITEMS_PER_PAGE) -> list[Carton]:
    pages: list[Carton] = []
    for carton in cartons:
        chunks = [carton.items[i:i + page_size] for i in range(0, len(carton.items), page_size)] or [[]]
        carton_total = decimal_text(sum((_qty_number(item.qty) for item in carton.items), Decimal(0)))
        for part, items in enumerate(chunks, start=1):
            display_carton = carton.carton if len(chunks) == 1 else _continued_carton_number(carton.carton, part)
            pages.append(Carton(
                carton=display_carton,
                or_no=carton.or_no,
                ref_no=carton.ref_no,
                gross_weight=carton.gross_weight,
                packaging_code=carton.packaging_code,
                items=list(items),
                show_total=part == len(chunks),
                total_qty=carton_total,
            ))
    return pages


def build_sublist_pdf(cartons: list[Carton]) -> bytes:
    output = BytesIO()
    c = canvas.Canvas(output, pagesize=A5, pageCompression=1)
    page_w, page_h = A5
    margin = 7 * mm
    content_w = page_w - 2 * margin
    for carton in paginate_cartons(cartons):
        item_count = len(carton.items)
        compact = item_count > 24
        y = page_h - (5 if compact else 7) * mm
        meta = [
            ("Carton #", carton.carton),
            ("OR #", carton.or_no.replace("_", " ")),
            ("Ref #", carton.ref_no),
            ("GW", f"{carton.gross_weight} KG" if carton.gross_weight and "kg" not in carton.gross_weight.lower() else carton.gross_weight),
        ]
        meta.append(("Packing Code #", carton.packaging_code))
        meta_font = 9 if compact else 10.5
        meta_leading = meta_font + 1.8
        label_style = ParagraphStyle(
            "meta-label", fontName="Helvetica-Bold", fontSize=meta_font,
            leading=meta_leading, alignment=0,
        )
        value_style = ParagraphStyle(
            "meta-value", fontName="Helvetica", fontSize=meta_font,
            leading=meta_leading, alignment=0, splitLongWords=True,
        )
        meta_data = [
            [Paragraph(label, label_style), Paragraph(value or "", value_style)]
            for label, value in meta
        ]
        meta_table = Table(meta_data, colWidths=[32 * mm, content_w - 32 * mm])
        meta_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (0, -1), 0),
            ("RIGHTPADDING", (0, 0), (0, -1), 3 * mm),
            ("LEFTPADDING", (1, 0), (1, -1), 0),
            ("RIGHTPADDING", (1, 0), (1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 1.2 * mm),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1.2 * mm),
        ]))
        _, meta_h = meta_table.wrap(content_w, page_h)
        meta_table.drawOn(c, margin, y - meta_h)
        y -= meta_h + (5 if compact else 8) * mm

        display_rows = max(15, len(carton.items))
        bottom_reserve = 13 * mm
        header_h = (5.5 if compact else 7) * mm
        available_h = y - bottom_reserve
        row_h = min(7.2 * mm, (available_h - header_h) / display_rows)
        if row_h < 2.5 * mm:
            raise ValueError(
                f"Carton {carton.carton} có {item_count} dòng, vượt khả năng hiển thị rõ trên một trang A5."
            )
        font_size = min(10, max(4.5, row_h / mm * 1.55))
        styles = ParagraphStyle(
            "cell", fontName="Helvetica", fontSize=font_size,
            # Keep the text box close to the glyph height so ReportLab can
            # center it vertically inside the full table row.
            leading=font_size + 1, alignment=TA_CENTER,
        )
        item_style = ParagraphStyle(
            "item-cell", parent=styles, alignment=0,
        )
        header_style = ParagraphStyle("head", parent=styles, fontName="Helvetica-Bold")
        data = [[Paragraph("Item No.", header_style), Paragraph("EAN", header_style), Paragraph("QTY", header_style)]]
        for item in carton.items:
            data.append([Paragraph(item.sku, item_style), Paragraph(item.ean, styles), Paragraph(item.qty, styles)])
        for _ in range(display_rows - len(carton.items)):
            data.append(["", "", ""])
        table = Table(data, colWidths=[55 * mm, 43 * mm, content_w - 98 * mm], rowHeights=[header_h] + [row_h] * display_rows)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e7e7e7")),
            ("GRID", (0, 0), (-1, -1), 0.65, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("ALIGN", (0, 1), (0, -1), "LEFT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 1.5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 1.5),
            ("LEFTPADDING", (0, 1), (0, -1), 3 * mm),
            ("TOPPADDING", (0, 0), (-1, -1), 1.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0.5),
        ]))
        table_h = header_h + row_h * display_rows
        table.wrapOn(c, content_w, table_h)
        table.drawOn(c, margin, y - table_h)
        if carton.show_total:
            total_text = carton.total_qty or decimal_text(
                sum((_qty_number(item.qty) for item in carton.items), Decimal(0))
            )
            c.setFont("Helvetica-Bold", 11)
            qty_left = margin + 98 * mm
            qty_width = content_w - 98 * mm
            c.drawCentredString(qty_left + qty_width / 2, y - table_h - 7 * mm, total_text)
        c.showPage()
    c.save()
    return output.getvalue()
